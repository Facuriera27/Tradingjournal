import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

# Load the original CSV data
df = pd.read_csv('trades.csv')

# Convert date columns to datetime
df['Open Time'] = pd.to_datetime(df['Open Time'], dayfirst=True, errors='coerce')
df['Close Time'] = pd.to_datetime(df['Close Time'], dayfirst=True, errors='coerce')

# Filter out Break Even trades (Profit = 0)
df = df[df['Profit (USD)'] != 0]

# Calculate trade duration (in hours)
df['Duration (hours)'] = (df['Close Time'] - df['Open Time']).dt.total_seconds() / 3600

# Calculate win/loss
df['Result'] = df['Profit (USD)'].apply(lambda x: 'Win' if x > 0 else 'Loss')

# Add day of week
df['Day of Week'] = df['Open Time'].dt.day_name()

# Add month (in English for consistency)
df['Month'] = df['Open Time'].dt.month_name()

# Add time of day categories
def get_time_of_day(hour):
    if 0 <= hour < 6:
        return 'Night'
    elif 6 <= hour < 12:
        return 'Morning'
    elif 12 <= hour < 18:
        return 'Afternoon'
    else:
        return 'Evening'

df['Time of Day'] = df['Open Time'].dt.hour.apply(get_time_of_day)

# Calculate risk-reward ratio
def calculate_rr(row):
    if pd.notna(row['Take Profit']) and pd.notna(row['Stop Loss']) and pd.notna(row['Open Price']):
        if row['Side'] == 'BUY':
            risk = row['Open Price'] - row['Stop Loss']
            reward = row['Take Profit'] - row['Open Price']
        else:  # SELL
            risk = row['Stop Loss'] - row['Open Price']
            reward = row['Open Price'] - row['Take Profit']
        
        if risk != 0:
            return round(reward / risk, 2)
    return None

df['Risk-Reward'] = df.apply(calculate_rr, axis=1)

# Select only the columns we want to keep
selected_columns = [
    'Symbol', 'Side', 'Size', 'Take Profit', 'Stop Loss', 
    'Profit (USD)', 'Duration (hours)', 'Result', 
    'Day of Week', 'Month', 'Time of Day', 'Risk-Reward'
]
df = df[selected_columns]

# Create summary sheet
summary_data = {
    'Metric': ['Total Trades', 'Profitable Trades', 'Losing Trades', 
               'Total Profit (USD)', 'Average Profit per Trade', 'Win Rate', 'Profit Factor'],
    'Value': [
        len(df),
        len(df[df['Result'] == 'Win']),
        len(df[df['Result'] == 'Loss']),
        df['Profit (USD)'].sum(),
        df['Profit (USD)'].mean(),
        len(df[df['Result'] == 'Win']) / len(df),
        abs(df[df['Result'] == 'Win']['Profit (USD)'].sum()) / abs(df[df['Result'] == 'Loss']['Profit (USD)'].sum())
    ]
}

summary_df = pd.DataFrame(summary_data)

# Create performance by symbol sheet
symbol_stats = df.groupby('Symbol').agg(
    Total_Profit=('Profit (USD)', 'sum'),
    Trade_Count=('Profit (USD)', 'count'),
    Avg_Profit=('Profit (USD)', 'mean'),
    Win_Rate=('Result', lambda x: (x == 'Win').mean())
).reset_index()

# Create performance by time of day sheet
time_stats = df.groupby('Time of Day').agg(
    Total_Profit=('Profit (USD)', 'sum'),
    Trade_Count=('Profit (USD)', 'count'),
    Avg_Profit=('Profit (USD)', 'mean'),
    Win_Rate=('Result', lambda x: (x == 'Win').mean())
).reset_index()

# Create performance by day of week sheet
day_stats = df.groupby('Day of Week').agg(
    Total_Profit=('Profit (USD)', 'sum'),
    Trade_Count=('Profit (USD)', 'count'),
    Avg_Profit=('Profit (USD)', 'mean'),
    Win_Rate=('Result', lambda x: (x == 'Win').mean())
).reset_index()

# Define color palette for months (12 distinct colors)
month_colors = {
    'January': 'FF9999',
    'February': '99FF99',
    'March': '9999FF',
    'April': 'FFFF99',
    'May': 'FF99FF',
    'June': '99FFFF',
    'July': 'FFCC99',
    'August': 'CCFF99',
    'September': '99CCFF',
    'October': 'FF99CC',
    'November': 'CC99FF',
    'December': '99FFCC'
}

# Export to Excel with formatting
with pd.ExcelWriter('trading_journal_month_colors.xlsx', engine='openpyxl') as writer:
    # Export data sheets
    df.to_excel(writer, sheet_name='All Trades', index=False)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    symbol_stats.to_excel(writer, sheet_name='By Symbol', index=False)
    time_stats.to_excel(writer, sheet_name='By Time of Day', index=False)
    day_stats.to_excel(writer, sheet_name='By Day of Week', index=False)
    
    # Get the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['All Trades']
    
    # Define colors for Result column
    green_fill = PatternFill(start_color='00CCFFCC', end_color='00CCFFCC', fill_type='solid')
    red_fill = PatternFill(start_color='00FFCCCC', end_color='00FFCCCC', fill_type='solid')
    
    # Create fills for each month
    month_fills = {month: PatternFill(start_color=color, end_color=color, fill_type='solid') 
                  for month, color in month_colors.items()}
    
    # Apply color formatting
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        # Color Result column (H)
        result_cell = row[7]  # Column H (0-based index 7)
        if result_cell.value == 'Win':
            result_cell.fill = green_fill
        elif result_cell.value == 'Loss':
            result_cell.fill = red_fill
        
        # Color Month column (J)
        month_cell = row[9]  # Column J (0-based index 9)
        if month_cell.value in month_fills:
            month_cell.fill = month_fills[month_cell.value]
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

print("Trading journal with month colors exported to 'trading_journal_month_colors.xlsx'")